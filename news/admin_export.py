# news/admin_export.py
from django.utils.encoding import force_str
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from .models import News


def safe_value(value):
    if value is None:
        return ''
    if isinstance(value, (int, float, bool)):
        return value
    if hasattr(value, 'strftime'):
        return value.isoformat()
    return str(force_str(value))


def generate_news_centric_xlsx(fields):
    """
    Экспорт от News с LEFT JOIN к Category, Author и Vote.
    Используем prefetch_related и аннотации для связей с Vote.

    """

    # Получаем все новости с предзагрузкой связей
    news_queryset = News.objects.select_related('category', 'author').prefetch_related('votes__user')

    all_rows = []
    for news in news_queryset:
        votes = list(news.votes.select_related('user'))  # голоса по этой новости
        if not votes:
            # Новость без голосов — одна строка
            row = _build_row(news, None, fields)
            all_rows.append(row)
        else:
            # Новость с голосами — по строке на каждый голос
            for vote in votes:
                row = _build_row(news, vote, fields)
                all_rows.append(row)

    # Собираем все возможные колонки
    all_columns = []
    for row in all_rows:
        for key in row.keys():
            if key not in all_columns:
                all_columns.append(key)

    # Создаём Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "News Report"

    # Заголовки
    for col_num, col in enumerate(all_columns, 1):
        ws[get_column_letter(col_num) + '1'] = col

    # Данные
    for row_num, row in enumerate(all_rows, 2):
        for col_num, col in enumerate(all_columns, 1):
            value = row.get(col)
            ws[get_column_letter(col_num) + str(row_num)] = safe_value(value)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _build_row(news, vote, fields):
    """Собирает одну строку отчёта."""
    row = {}

    # Поля News
    if 'News_id' in fields:
        row['News_id'] = news.id
    if 'News_title' in fields:
        row['News_title'] = news.title
    if 'News_content' in fields:
        row['News_content'] = news.content
    if 'News_is_published' in fields:
        row['News_is_published'] = news.is_published
    if 'News_created_at' in fields:
        row['News_created_at'] = news.created_at
    if 'News_updated_at' in fields:
        row['News_updated_at'] = news.updated_at

    # Поля Category
    if news.category:
        if 'Category_id' in fields:
            row['Category_id'] = news.category.id
        if 'Category_name' in fields:
            row['Category_name'] = news.category.name
        if 'Category_slug' in fields:
            row['Category_slug'] = news.category.slug

    # Поля Author
    if news.author:
        if 'Author_id' in fields:
            row['Author_id'] = news.author.id
        if 'Author_username' in fields:
            row['Author_username'] = news.author.username
        if 'Author_email' in fields:
            row['Author_email'] = news.author.email

    # Поля Vote и Voter
    if vote:
        if 'Vote_id' in fields:
            row['Vote_id'] = vote.id
        if 'Vote_value' in fields:
            row['Vote_value'] = vote.value
        if 'Vote_created_at' in fields:
            row['Vote_created_at'] = vote.created_at
        if 'Vote_updated_at' in fields:
            row['Vote_updated_at'] = vote.updated_at

        if vote.user:
            if 'Voter_id' in fields:
                row['Voter_id'] = vote.user.id
            if 'Voter_username' in fields:
                row['Voter_username'] = vote.user.username
            if 'Voter_email' in fields:
                row['Voter_email'] = vote.user.email
    else:
        # Заполняем пустые значения для полей голоса, если они выбраны
        for f in fields:
            if f.startswith('Vote_') or f.startswith('Voter_'):
                row[f] = ''

    return row